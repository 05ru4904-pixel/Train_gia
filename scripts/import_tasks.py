"""Массовый импорт заданий из таблицы (Excel или CSV).

Через админ-бота задания добавляются по одному, и на тысяче штук это нереально:
у Telegram лимит 4096 символов на сообщение. Этот скрипт читает таблицу целиком,
проверяет каждую строку и заливает всё разом.

    python scripts/import_tasks.py задания.xlsx --dry-run   # только проверить
    python scripts/import_tasks.py задания.xlsx             # проверить и залить
    python scripts/import_tasks.py --template               # создать шаблон таблицы

Ожидаемые колонки (порядок неважен, регистр и лишние пробелы игнорируются):

    номер | вопрос | вариант 1 | вариант 2 | ... | ответ

Пустые ячейки вариантов пропускаются — поэтому в одной таблице спокойно уживаются
задания с четырьмя, пятью и шестью вариантами. В колонке «ответ» пишутся буквы
правильных вариантов через запятую: «А» или «А, В, Д». Цифры («1, 3») тоже поймёт.
"""
import argparse
import asyncio
import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from core.parser import ParsedTask, ParseError, parse_answer  # noqa: E402
from core.tasks_meta import LAST_TASK, letter  # noqa: E402
from db import crud  # noqa: E402
from db.database import SessionMaker, dispose_db, init_db  # noqa: E402

# Как могут называться колонки. Сравнение идёт по нижнему регистру без пробелов.
ALIASES_NUMBER = {"номер", "№", "n", "number", "задание", "номерзадания"}
ALIASES_TEXT = {"вопрос", "текст", "текстзадания", "условие", "question", "text"}
ALIASES_ANSWER = {"ответ", "правильныйответ", "ответы", "answer", "correct"}
RE_OPTION_HEADER = re.compile(r"^(?:вариант|ответ|option|opt)?\s*([1-9]\d?|[а-яa-z])$", re.I)

TEMPLATE_ROWS = [
    ["номер", "вопрос", "вариант 1", "вариант 2", "вариант 3", "вариант 4", "вариант 5", "ответ"],
    ["4", "В каком слове верно поставлено ударение?", "звонИт", "звОнит", "позвОнит", "тОрты", "", "А, Г"],
    ["5", "В каком предложении верно употреблён пароним?", "надеть пальто", "одеть пальто", "", "", "", "А"],
]


@dataclass
class RowError:
    row: int
    message: str


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).strip().lower()


def read_table(path: Path, sheet: str | None) -> list[list[str]]:
    """Возвращает таблицу как список строк, включая заголовок."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path, sheet)
    if suffix in {".csv", ".txt", ".tsv"}:
        return _read_csv(path)
    raise SystemExit(f"Не знаю, как читать {suffix}. Поддерживаются .xlsx и .csv")


def _read_xlsx(path: Path, sheet: str | None) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "Для .xlsx нужна библиотека openpyxl. Установите её:\n"
            "    pip install openpyxl\n"
            "Либо сохраните таблицу как CSV — его скрипт читает без дополнительных библиотек."
        ) from None

    book = load_workbook(path, read_only=True, data_only=True)
    page = book[sheet] if sheet else book.active
    rows = []
    for raw in page.iter_rows(values_only=True):
        rows.append(["" if cell is None else str(cell).strip() for cell in raw])
    book.close()
    return rows


def _read_csv(path: Path) -> list[list[str]]:
    # Excel в русской локали сохраняет CSV в cp1251 и с точкой с запятой —
    # пробуем оба варианта кодировки, разделитель определяем автоматически.
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise SystemExit("Не удалось прочитать файл: неизвестная кодировка. Пересохраните в UTF-8.")

    sample = text[:4096]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    return [[cell.strip() for cell in row] for row in csv.reader(text.splitlines(), delimiter=delimiter)]


def locate_columns(header: list[str]) -> tuple[int, int, list[int], int]:
    """Находит индексы колонок: номер, вопрос, варианты, ответ."""
    number_at = text_at = answer_at = -1
    option_at: list[int] = []

    for index, raw in enumerate(header):
        name = normalize_header(raw)
        if not name:
            continue
        if number_at < 0 and name in ALIASES_NUMBER:
            number_at = index
        elif text_at < 0 and name in ALIASES_TEXT:
            text_at = index
        elif name in ALIASES_ANSWER and index == len(header) - 1:
            answer_at = index
        elif RE_OPTION_HEADER.match(name) or name.startswith("вариант"):
            option_at.append(index)
        elif answer_at < 0 and name in ALIASES_ANSWER:
            answer_at = index

    missing = []
    if number_at < 0:
        missing.append("номер")
    if text_at < 0:
        missing.append("вопрос")
    if answer_at < 0:
        missing.append("ответ")
    if len(option_at) < 2:
        missing.append("минимум две колонки вариантов")
    if missing:
        raise SystemExit(
            "В заголовке таблицы не нашёл: " + ", ".join(missing) + ".\n"
            f"Прочитанный заголовок: {header}\n"
            "Создать образец правильной таблицы: python scripts/import_tasks.py --template"
        )
    return number_at, text_at, option_at, answer_at


def parse_row(row: list[str], columns, row_number: int) -> ParsedTask:
    number_at, text_at, option_at, answer_at = columns

    def cell(index: int) -> str:
        return row[index].strip() if index < len(row) else ""

    raw_number = cell(number_at)
    if not raw_number:
        raise ParseError("не указан номер задания")
    # Excel часто отдаёт числа как «4.0» — приводим к целому.
    try:
        number = int(float(raw_number.replace(",", ".")))
    except ValueError:
        raise ParseError(f"номер «{raw_number}» не похож на число") from None
    if not 1 <= number <= LAST_TASK:
        raise ParseError(f"номер должен быть от 1 до {LAST_TASK}, а указан {number}")

    text = cell(text_at)
    if not text:
        raise ParseError("пустой текст вопроса")

    options = [cell(i) for i in option_at]
    # Пустые ячейки в конце — это просто «вариантов меньше». А дырка в середине
    # почти всегда означает съехавшие данные, поэтому о ней сообщаем.
    while options and not options[-1]:
        options.pop()
    if len(options) < 2:
        raise ParseError("нужно минимум два варианта ответа")
    for position, value in enumerate(options):
        if not value:
            raise ParseError(
                f"вариант {letter(position)} пустой, хотя дальше варианты есть — "
                "похоже, данные в строке съехали"
            )

    correct = parse_answer(cell(answer_at), len(options))
    return ParsedTask(number=number, text=text, options=options, correct=correct)


def parse_table(rows: list[list[str]]) -> tuple[list[ParsedTask], list[RowError]]:
    if not rows:
        raise SystemExit("Файл пустой.")

    columns = locate_columns(rows[0])
    tasks: list[ParsedTask] = []
    errors: list[RowError] = []

    for offset, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue  # пустая строка-разделитель
        try:
            tasks.append(parse_row(row, columns, offset))
        except ParseError as exc:
            errors.append(RowError(row=offset, message=str(exc)))
    return tasks, errors


def write_template(path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        csv.writer(handle, delimiter=";").writerows(TEMPLATE_ROWS)
    print(f"Шаблон записан: {path}")
    print("Откройте его в Excel, заполните и сохраните — можно как .xlsx, можно как .csv.")


async def do_import(tasks: list[ParsedTask], skip_duplicates: bool) -> None:
    await init_db()
    created = 0
    skipped = 0
    async with SessionMaker() as db:
        known: set[tuple[int, str]] = set()
        if skip_duplicates:
            for number in {task.number for task in tasks}:
                for existing in await crud.list_tasks(db, number, limit=100_000):
                    known.add((existing.number, existing.text.strip()))

        for task in tasks:
            key = (task.number, task.text.strip())
            if skip_duplicates and key in known:
                skipped += 1
                continue
            await crud.create_task(db, task)
            known.add(key)
            created += 1
            if created % 100 == 0:
                print(f"  залито {created}...")

        total = await crud.total_tasks_count(db)

    print(f"\nДобавлено заданий: {created}")
    if skipped:
        print(f"Пропущено дублей: {skipped}")
    print(f"Всего в базе: {total}")


async def main() -> None:
    parser = argparse.ArgumentParser(description="Импорт заданий из таблицы")
    parser.add_argument("file", nargs="?", help="файл .xlsx или .csv")
    parser.add_argument("--sheet", help="имя листа в книге Excel")
    parser.add_argument("--dry-run", action="store_true", help="только проверить, ничего не записывать")
    parser.add_argument("--skip-invalid", action="store_true", help="залить корректные строки, пропустив битые")
    parser.add_argument("--skip-duplicates", action="store_true", help="не добавлять задания с таким же текстом")
    parser.add_argument("--template", action="store_true", help="создать образец таблицы и выйти")
    args = parser.parse_args()

    if args.template:
        write_template(Path("шаблон_заданий.csv"))
        return
    if not args.file:
        parser.error("укажите файл или --template")

    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f"Файл не найден: {path}")

    rows = read_table(path, args.sheet)
    tasks, errors = parse_table(rows)

    print(f"Строк в файле: {len(rows) - 1}")
    print(f"Разобрано заданий: {len(tasks)}")

    by_number: dict[int, int] = {}
    for task in tasks:
        by_number[task.number] = by_number.get(task.number, 0) + 1
    if by_number:
        spread = ", ".join(f"№{n}: {by_number[n]}" for n in sorted(by_number))
        print(f"По номерам — {spread}")

    if errors:
        print(f"\nСтрок с ошибками: {len(errors)}")
        for error in errors[:25]:
            print(f"  строка {error.row}: {error.message}")
        if len(errors) > 25:
            print(f"  ... и ещё {len(errors) - 25}")
        if not args.skip_invalid:
            print(
                "\nНичего не залито. Исправьте строки выше и запустите снова "
                "либо добавьте --skip-invalid, чтобы залить остальные."
            )
            return

    if args.dry_run:
        print("\nПроверка завершена, база не тронута (--dry-run).")
        return
    if not tasks:
        print("\nЗаливать нечего.")
        return

    try:
        await do_import(tasks, args.skip_duplicates)
    finally:
        await dispose_db()


if __name__ == "__main__":
    asyncio.run(main())
